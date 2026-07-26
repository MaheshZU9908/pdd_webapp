# automation/utils/excel_generator.py
"""
Enterprise Excel Report Generator using openpyxl.
Generates multi-sheet workbook containing:
Sheet 1: Executed Test Cases (Test ID, Module, Test Name, Status, Execution Time, Priority)
Sheet 2: Passed Tests
Sheet 3: Failed Tests
Sheet 4: Skipped Tests
Sheet 5: Execution Metrics
Sheet 6: Defect Summary
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReportGenerator:
    def __init__(self, filename, title_banner="PathoAI Automation Test Execution Report"):
        self.filename = filename
        self.title_banner = title_banner

    def generate(self, test_cases):
        os.makedirs(os.path.dirname(self.filename), exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        total_count = len(test_cases)
        passed_cases = [tc for tc in test_cases if tc.get("status") == "PASS"]
        failed_cases = [tc for tc in test_cases if tc.get("status") == "FAIL"]
        skipped_cases = [tc for tc in test_cases if tc.get("status") == "SKIP"]

        # ── Sheet 1: Executed Test Cases ──
        ws1 = wb.create_sheet(title="Executed Test Cases")
        self._build_test_cases_sheet(ws1, "ALL EXECUTED TEST CASES", test_cases)

        # ── Sheet 2: Passed Tests ──
        ws2 = wb.create_sheet(title="Passed Tests")
        self._build_test_cases_sheet(ws2, "PASSED TEST CASES", passed_cases)

        # ── Sheet 3: Failed Tests ──
        ws3 = wb.create_sheet(title="Failed Tests")
        self._build_test_cases_sheet(ws3, "FAILED TEST CASES", failed_cases)

        # ── Sheet 4: Skipped Tests ──
        ws4 = wb.create_sheet(title="Skipped Tests")
        self._build_test_cases_sheet(ws4, "SKIPPED TEST CASES", skipped_cases)

        # ── Sheet 5: Execution Metrics ──
        ws5 = wb.create_sheet(title="Execution Metrics")
        self._build_metrics_sheet(ws5, total_count, len(passed_cases), len(failed_cases), len(skipped_cases), test_cases)

        # ── Sheet 6: Defect Summary ──
        ws6 = wb.create_sheet(title="Defect Summary")
        self._build_defect_sheet(ws6, failed_cases)

        wb.save(self.filename)
        return self.filename

    def _build_test_cases_sheet(self, ws, banner_text, cases):
        ws.views.sheetView[0].showGridLines = True
        
        # Banner Header
        ws.merge_cells("A1:F2")
        b = ws["A1"]
        b.value = f"{self.title_banner} — {banner_text}"
        b.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        b.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        b.alignment = Alignment(horizontal="center", vertical="center")

        headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority"]
        ws.append([]) # Row 3 blank
        ws.append(headers) # Row 4 headers

        hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        hdr_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center" if col_idx in [1,4,5,6] else "left", vertical="center")

        pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        skip_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        pass_font = Font(name="Segoe UI", size=10, bold=True, color="15803D")
        fail_font = Font(name="Segoe UI", size=10, bold=True, color="B91C1C")
        skip_font = Font(name="Segoe UI", size=10, bold=True, color="B45309")

        thin_border = Border(
            left=Side(style='thin', color="E2E8F0"),
            right=Side(style='thin', color="E2E8F0"),
            top=Side(style='thin', color="E2E8F0"),
            bottom=Side(style='thin', color="E2E8F0")
        )

        for tc in cases:
            st = tc.get("status", "PASS")
            row = [
                tc.get("test_id", ""),
                tc.get("module", ""),
                tc.get("title", ""),
                st,
                tc.get("duration", 0.0),
                tc.get("priority", "P2")
            ]
            ws.append(row)
            r_idx = ws.max_row
            
            # Formatting
            for c_idx in range(1, 7):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                cell.font = Font(name="Segoe UI", size=10)
                cell.alignment = Alignment(vertical="center", horizontal="center" if c_idx in [1,4,5,6] else "left")

            # Status cell styling
            st_cell = ws.cell(row=r_idx, column=4)
            if st == "PASS":
                st_cell.fill = pass_fill; st_cell.font = pass_font
            elif st == "FAIL":
                st_cell.fill = fail_fill; st_cell.font = fail_font
            else:
                st_cell.fill = skip_fill; st_cell.font = skip_font

        # Auto-fit widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    def _build_metrics_sheet(self, ws, total, passed, failed, skipped, cases):
        ws.views.sheetView[0].showGridLines = True
        ws.merge_cells("A1:D2")
        b = ws["A1"]
        b.value = f"{self.title_banner} — EXECUTION METRICS SUMMARY"
        b.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        b.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        b.alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])
        ws.append(["Metric", "Count / Value", "Percentage", "Target SLA"])
        
        pass_rate = (passed / total * 100) if total > 0 else 0
        total_duration = sum(tc.get("duration", 0) for tc in cases)

        metrics_data = [
            ("Total Test Cases", total, "100%", "300+ Required"),
            ("Passed Test Cases", passed, f"{(passed/total*100):.1f}%" if total else "0%", "≥ 95.0% Pass"),
            ("Failed Test Cases", failed, f"{(failed/total*100):.1f}%" if total else "0%", "< 5.0% Fail"),
            ("Skipped Test Cases", skipped, f"{(skipped/total*100):.1f}%" if total else "0%", "0% Skipped"),
            ("Overall Pass Percentage", f"{pass_rate:.1f}%", f"{pass_rate:.1f}%", "≥ 95.0% Goal"),
            ("Total Execution Time", f"{total_duration:.2f}s", "N/A", "< 60.0s SLA")
        ]

        hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        hdr_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

        for c_idx in range(1, 5):
            cell = ws.cell(row=4, column=c_idx)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for m in metrics_data:
            ws.append(list(m))
            r = ws.max_row
            for c_idx in range(1, 5):
                cell = ws.cell(row=r, column=c_idx)
                cell.font = Font(name="Segoe UI", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 30

    def _build_defect_sheet(self, ws, failed_cases):
        ws.views.sheetView[0].showGridLines = True
        ws.merge_cells("A1:E2")
        b = ws["A1"]
        b.value = f"{self.title_banner} — DEFECT SUMMARY LOG"
        b.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        b.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        b.alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])
        ws.append(["Test ID", "Module", "Failure Description", "Severity", "Action / Owner"])
        
        hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        hdr_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

        for c_idx in range(1, 6):
            cell = ws.cell(row=4, column=c_idx)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if not failed_cases:
            ws.append(["N/A", "All Modules", "Zero Defects Observed — All test cases passed!", "None", "QA Architect Approved"])
        else:
            for tc in failed_cases:
                ws.append([tc.get("test_id"), tc.get("module"), tc.get("actual"), "High", "DevOps / SDET Team"])

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 30
